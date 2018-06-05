from pathlib import Path, PurePosixPath
from abc import ABCMeta, abstractmethod
from csv import DictReader as CSVDictReader
from openpyxl import load_workbook, utils
from datetime import datetime, date
from validator import Validator


class FileTypeAbstract(metaclass=ABCMeta):
    def __init__(self):
        self.filename = None
        self.doc = None
        self.validator = Validator()

    def file_exist(self):
        """
        Checks if the current file exists
        """
        if self.filename.exists():
            return True
        else:
            return False

    def template_method(self):
        self.open()
        result = self.read()
        self.close()
        return result

    @abstractmethod
    def open(self):
        pass

    @abstractmethod
    def close(self):
        pass

    @abstractmethod
    def read(self):
        pass


class FileTypeXLSX(FileTypeAbstract):

    def open(self):
        workbook = load_workbook(self.filename)
        first_sheet = workbook.sheetnames[0]
        worksheet = workbook[first_sheet]
        self.doc = worksheet

    def close(self):
        pass

    def read(self):
        """
        Return dictionary with key => value pairs
        :param worksheet is the file where the values exist
        """
        data = dict()
        empno = 0
        keys = []
        a_row = 0
        try:
            for row in self.doc.iter_rows():
                record = dict()
                row_num = 0
                for cell in row:
                    a_row = cell.row
                    if 1 == a_row:
                        keys.append(cell.value)
                    else:
                        valid = cell.value
                        if isinstance(cell.value, datetime):
                            valid = Validator.xlsx_date(cell.value)
                        record[keys[row_num]] = valid
                    row_num += 1
                if a_row > 1:
                    data[empno] = record
                empno += 1
            result = data
            return result
        except PermissionError:
            print("Sorry, you don't have enough permissions to access this file")


class FileTypeCSV(FileTypeAbstract):

    def __init__(self):
        super().__init__()
        self.CSV = None

    def open(self):
        with open(self.filename, 'r') as f:
            self.CSV = f
            self.doc = CSVDictReader(f)

    def close(self):
        self.CSV.close()

    def read(self):
        """
        Return dictionary with key => value pairs
        :param filename is the file where the values exist
        >>> a=FileTypeCSV()
        >>> FileTypeCSV.read(a,graph_dgraph_data
        """
        try:
            data = dict()
            empno = 0
            for row in self.doc:
                record = dict()
                for key in row:
                    record[key] = row.get(key)
                data[empno] = record
                empno += 1
            result = data
            return result
        except TypeError:
            print("Error!!")


class FileTypeTXT(FileTypeAbstract):

    def open(self):
        with open(self.filename, 'r') as f:
            contents = f.readlines()
            self.doc = contents

    def close(self):
        pass

    def read(self):
        """
        Return dictionary with key => value pairs
        # :param filename is the file where the values exist
        """
        data = dict()
        empno = 0
        try:
            for line in self.doc:
                rows = line.split(";")
                dictionary = dict()
                for row in rows:
                    if len(row.split("=")) == 2:
                        key = row.split("=")[0]
                        value = row.split("=")[1]
                        value = value.rstrip('\n')
                        dictionary[key] = value
                        data[empno] = dictionary
                    else:
                        print("File error")
                        raise ValueError
                empno += 1
            result = self.validator.save_dict(data)
            return result

        except Exception as e:
            print(e)
